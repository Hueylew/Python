#!/usr/bin/env python3

from matplotlib import colors
import openpyxl
import os
import xlsxwriter
import pandas as pd
import numpy as np
import glob
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re
import fpdf
from fpdf import FPDF

def load_workbook_range(range_string, ws):
	col_start, col_end = re.findall("[A-Z]+", range_string)
	
	data_rows = []
	for row in ws[range_string]:
		data_rows.append([cell.value for cell in row])
		
	return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

# Read all spreadsheets in the named folder below
all_data = pd.DataFrame()
index = 0
for f in glob.glob('/Users/adamlewis/Documents/Work/Business Consultancy/Skills and Development/Combined Skills Profiles 2021/*.xlsx'):
	index += 1
	# Get the persons name
	
	from openpyxl import load_workbook
	wb = load_workbook(f, data_only=True)
	sh = wb["Details"]
	PersonName = sh["B2"].value
	PersonPos = sh["B3"].value
	PersonScore = sh["G22"].value
	PersonTarget = sh["G23"].value
	PersonVariance = sh["G24"].value
	PersonRole = sh["B3"].value
	WorkIn = sh["B4"].value
	Location = sh["B5"].value
	YearsGI = sh["B6"].value
	YearsBA = sh["B7"].value
	YearsSSP = sh["B8"].value
	
	#Create the data frame
	from itertools import islice
	ws = wb["ACT"]
	df_BA = load_workbook_range('A1:C77', ws)
	
	# Create and Populate the score column based on answer provided
	df_BA = df_BA.fillna(value=0)
	df_BA["C"].replace({"Basic": 1}, inplace=True)
	df_BA["C"].replace({"Intermediate": 2}, inplace=True)
	df_BA["C"].replace({"Advanced": 3}, inplace=True)
	df_BA["C"].replace({"SME": 4}, inplace=True)
	df_BA = df_BA.rename(columns={'A': 'Category'})
	df_BA = df_BA.rename(columns={'B': 'Topic'})
	df_BA = df_BA.rename(columns={'C': 'Skill Level'})
	df_BA.at[2:7,'Category']='Analytical thinking and problem solving'
	df_BA.at[9:11,'Category']='Communication skills'
	df_BA.at[13:15,'Category']='Interaction skills'
	df_BA.at[17:21,'Category']='KCM Author'
	df_BA.at[23:29,'Category']='Pure In-House Document production'
	df_BA.at[31:40,'Category']='Pre Rules Author'
	df_BA.at[42:50,'Category']='Pure Product Dev - Back office'
	df_BA.at[52:59,'Category']='Pure Product Dev - PB2 Basic'
	df_BA.at[61:64,'Category']='Pure Product Dev - PB2 Advanced'
	df_BA.at[66:70,'Category']='Programming languages'
	df_BA.at[72,'Category']='Integration'
	df_BA.at[74:75,'Category']='BDX'

	all_data = df_BA
		
all_data.fillna('', inplace=True)
new_header = all_data.iloc[0] 
all_data = all_data[1:] 
all_data.columns = new_header
print(all_data)

groupCategory = all_data.groupby('Category')['Skill Level'].mean().sort_values(ascending=False)
print(groupCategory)
groupCategory.plot.bar(x="Category", y='Skill Level', rot=90, color='y', xlabel = 'Category', ylabel = 'Score', title="Score by category")
plt.tick_params(axis='x', labelsize=8)
plt.tick_params(axis='x', pad=-250)
plt.savefig('ACT Average Scores.png')
# plt.show(block=True)

document = fpdf.FPDF()

document.set_font('Helvetica', style='B', size=20)
document.set_text_color(19, 83, 173)
document.add_page()
document.cell(60)
document.cell(70, 10, PersonName, 1, 0, "C")
document.ln(20)
document.set_font('Helvetica', style='', size=12)
document.set_text_color(0)
document.cell(0, h=5, txt = 'Role: ' + PersonRole + ' working in ' + WorkIn)
document.ln(5)
document.cell(0, h=5, txt = 'Location: ' + Location)
document.ln(5)
document.cell(0, h=5, txt = 'Years at SSP: ' + str(YearsSSP))
document.ln(5)

# Add in ACT Scores Plot Image
document.image('ACT Average Scores.png', w=150, h=100)
document.ln(30)

#Save the PDF Document
document.output(PersonName + ' - Skills Profile Report.pdf')