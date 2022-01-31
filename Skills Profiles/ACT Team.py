#!/usr/bin/env python3

import openpyxl
import os
import xlsxwriter
import pandas as pd
import numpy as np
import glob
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re

def load_workbook_range(range_string, ws):
	col_start, col_end = re.findall("[A-Z]+", range_string)
	
	data_rows = []
	for row in ws[range_string]:
		data_rows.append([cell.value for cell in row])
		
	return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

# Read all spreadsheets in the named folder below
all_data = pd.DataFrame()
index = 0
for f in glob.glob('/Users/adamlewis/Documents/Work/Business Consultancy/Skills and Development/Combined Skills Profiles 2022/*.xlsx'):
	index += 1
	# Get the persons name
	
	from openpyxl import load_workbook
	wb = load_workbook(f, data_only=True)
	sh = wb["Details"]
	PersonName = sh["B2"].value
	PersonPos = sh["B3"].value
	PersonScore = sh["G22"].value
	PersonTarget = sh["G23"].value
	PersonVariance = sh["G24"].value
	PersonRole = sh["B3"].value
	WorkIn = sh["B4"].value
	Location = sh["B5"].value
	YearsGI = sh["B6"].value
	YearsBA = sh["B7"].value
	YearsSSP = sh["B8"].value
	
	#Create the data frame
	from itertools import islice
	ws = wb["ACT"]
	df_BA = load_workbook_range('A1:C77', ws)
	
	# Create and Populate the score column based on answer provided
	df_BA = df_BA.fillna(value=0)
	df_BA["C"].replace({"Basic": 1}, inplace=True)
	df_BA["C"].replace({"Intermediate": 2}, inplace=True)
	df_BA["C"].replace({"Advanced": 3}, inplace=True)
	df_BA["C"].replace({"SME": 4}, inplace=True)
	df_BA = df_BA.rename(columns={'A': 'Category'})
	df_BA = df_BA.rename(columns={'B': 'Topic'})
	df_BA = df_BA.rename(columns={'C': PersonName})
	df_BA = df_BA.append({'Category':'Score', 'Topic':'Score', PersonName : PersonScore}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Target', 'Topic':'Target', PersonName : PersonTarget}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Variance', 'Topic':'Variance', PersonName : PersonVariance}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Role', 'Topic':'Role', PersonName : PersonRole}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Work in', 'Topic':'Work in', PersonName : WorkIn}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Location', 'Topic':'Location', PersonName : Location}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Years GI', 'Topic':'Years GI', PersonName : YearsGI}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Years BA', 'Topic':'Years BA', PersonName : YearsBA}, ignore_index=True)
	df_BA = df_BA.append({'Category':'Years SSP', 'Topic':'Years SSP', PersonName : YearsSSP}, ignore_index=True)
	
	#merge all dataframes together into one large dataframe
	if index == 1:
		all_data = df_BA
	else:
		all_data = pd.merge(all_data, df_BA)
		
all_data.fillna('', inplace=True)
print(all_data)

# save to excel file
book = load_workbook('/Users/adamlewis/BC Team Overall.xlsx')
writer = pd.ExcelWriter('/Users/adamlewis/BC Team Overall.xlsx', engine='openpyxl') 
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

all_data.to_excel(writer, "ACT", index = False)

writer.save()