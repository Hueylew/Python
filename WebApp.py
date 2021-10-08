from logging import Filter
from matplotlib import colors
import openpyxl
import os
from openpyxl.utils.cell import column_index_from_string
import xlsxwriter
import altair as alt
from altair import *
import pandas as pd
import streamlit as st
import plotly.express as px
from PIL import Image
import numpy as np
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re


def load_workbook_range(range_string, ws):
	col_start, col_end = re.findall("[A-Z]+", range_string)
	
	data_rows = []
	for row in ws[range_string]:
		data_rows.append([cell.value for cell in row])
		
	return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


# Create sidebar
st.set_page_config(page_title='Skills Profiles')
st.sidebar.subheader('Dashboard Options')
display = ['Adam Lewis','Alex Bandini','Andrew Wilcock','Ankit Singh','Arnold Kaswa','Badimu Kazadi','Billy May','Charles Chabalala','Claude Seabi',
'Daniso Mushaike','David Newson','Doctor Bokisi', 'Donna Cochrane', 'Gemal Dabab', 'Grant Smalley', 'Jalal Saleem', 'Kamogelo Matiou','Khanyisa Makhubela',
'Michelle Padyachee','Mo Harmen', 'Muhammad Adam','Palesa Khantsi','Phuluso Ramulifho','Prevesh Kuni','Rachna Kumar','Ratidzo Marowatsanga','Richard Heath',
'Saumitra Bhatnagar','Stephanie Davis','Steve Deas','Sucheta Mohan','Tebello Khesa','Teboho Monareng','Terry Ndou','Thinh Nguyen','Wing Chan']
result = st.sidebar.selectbox('Select team member', display)
businessAreas = ['ACT','Architecture','BA','Domain','IQH','Pure Ins','Select']
businessAreasResult = st.sidebar.selectbox('Business Area', businessAreas)


# Read the spreadsheet for the selected person from the sidebar above
all_data = pd.DataFrame()
f = ('/Users/adamlewis/Documents/Work/Business Consultancy/Skills and Development/Combined Skills Profiles 2021 copy/' + result + ' - Skills Profile.xlsx')

# Get personal information from the Details sheet
from openpyxl import load_workbook
wb = load_workbook(f, data_only=True)
sh = wb["Details"]
PersonName = sh["B2"].value
PersonPos = sh["B3"].value
PersonScore = sh["G22"].value
PersonTarget = sh["G23"].value
PersonVariance = sh["G24"].value
PersonRole = sh["B3"].value
WorkIn = sh["B4"].value
Location = sh["B5"].value
YearsGI = sh["B6"].value
YearsBA = sh["B7"].value
YearsSSP = sh["B8"].value

#Create the scores data frame
from itertools import islice
ws = wb["Details"]
df_Scores = load_workbook_range('B22:K24', ws)
df_Scores = df_Scores.rename(columns={'B': '','C': 'Select','D': 'Pure Ins', 'E': 'SSP Broker', 'F': 'IQH', 'G': 'ACT', 'H': 'Common Components', 'I': 'Domain', 'J': 'BA', 'K': 'Architecture'})
for col in df_Scores.describe().columns:
    df_Scores[col] = df_Scores[col].apply(lambda x: 1 if np.isnan(x) else round(x*100))

# Define the worksheets and ranges needed to power the business area choices
if businessAreasResult == 'ACT':
    #Create the data frame
    from itertools import islice
    ws = wb['ACT']
    df_BA = load_workbook_range('A1:C77', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["C"].replace({"Basic": 1}, inplace=True)
    df_BA["C"].replace({"Intermediate": 2}, inplace=True)
    df_BA["C"].replace({"Advanced": 3}, inplace=True)
    df_BA["C"].replace({"SME": 4}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Skill Level'})
    df_BA.at[2:7,'Category']='Analytical thinking and problem solving'
    df_BA.at[9:11,'Category']='Communication skills'
    df_BA.at[13:15,'Category']='Interaction skills'
    df_BA.at[17:21,'Category']='KCM Author'
    df_BA.at[23:29,'Category']='Pure In-House Document production'
    df_BA.at[31:40,'Category']='Pre Rules Author'
    df_BA.at[42:50,'Category']='Pure Product Dev - Back office'
    df_BA.at[52:59,'Category']='Pure Product Dev - PB2 Basic'
    df_BA.at[61:64,'Category']='Pure Product Dev - PB2 Advanced'
    df_BA.at[66:70,'Category']='Programming languages'
    df_BA.at[72,'Category']='Integration'
    df_BA.at[74:75,'Category']='BDX'
elif businessAreasResult == 'IQH':
    #Create the data frame
    from itertools import islice
    ws = wb["IQH"]
    df_BA = load_workbook_range('A1:C29', ws)
    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["C"].replace({"Basic": 1}, inplace=True)
    df_BA["C"].replace({"Intermediate": 2}, inplace=True)
    df_BA["C"].replace({"Advanced": 3}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Skill Level'})
    df_BA.at[5,'Category']='IQH BI Feed'
    df_BA.at[7,'Category']='IQH Working Data'
    df_BA.at[10:13,'Category']='IQH Data Enrichment Experian'
    df_BA.at[15:16,'Category']='IQH Data Enrichment Lexis Nexis'
    df_BA.at[26,'Category']='IQH Data Derivation'
    df_BA.at[27:28,'Category']='IQH Commission & Tax Calculation'
elif businessAreasResult == 'Select':
    #Create the data frame
    from itertools import islice
    ws = wb["SIaaS"]
    df_BA = load_workbook_range('A1:C213', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["C"].replace({"Basic": 1}, inplace=True)
    df_BA["C"].replace({"Intermediate": 2}, inplace=True)
    df_BA["C"].replace({"Advanced": 3}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Skill Level'})
    df_BA.at[1:18,'Category']='Configuration (including self sufficiency options)'
    df_BA.at[19:32,'Category']='Policy Admin (Excluding Renewals)'
    df_BA.at[33:46,'Category']='Policy Admin (Renewals)'
    df_BA.at[47:74,'Category']='Finance / Accounting'
    df_BA.at[75:87,'Category']='Claims'
    df_BA.at[88:92,'Category']='Payments Hub (Shared Component)'
    df_BA.at[93:101,'Category']='Clients / Contacts'
    df_BA.at[102:107,'Category']='Complaints'
    df_BA.at[108:121,'Category']='Product Building'
    df_BA.at[122,'Category']='Rating (Not in the actual product)'
    df_BA.at[123,'Category']='Front-end Product Building Tool'
    df_BA.at[124:140,'Category']='Reinsurance'
    df_BA.at[141:144,'Category']='Reporting'
    df_BA.at[145:146,'Category']='Reporting tool/dashboard'
    df_BA.at[147:171,'Category']='Front end processing'
    df_BA.at[172:173,'Category']='Work Management'
    df_BA.at[174:186,'Category']='Document Management'
    df_BA.at[187:193,'Category']='Kofax'
    df_BA.at[194:196,'Category']='Portfolio transfers'
    df_BA.at[197:200,'Category']='Integration to I/90'
    df_BA.at[201:210,'Category']='Batch processing (outside of the above categories)'
    df_BA.at[211:214,'Category']='Compliance'
elif businessAreasResult == 'Pure Ins':
    #Create the data frame
    from itertools import islice
    ws = wb["Pure Ins"]
    df_BA = load_workbook_range('A1:C170', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["C"].replace({"Basic": 1}, inplace=True)
    df_BA["C"].replace({"Intermediate": 2}, inplace=True)
    df_BA["C"].replace({"Advanced": 3}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Skill Level'})
    df_BA.at[1:20,'Category']='Configuration'
    df_BA.at[21:32,'Category']='Policy admin (excluding renewals)'
    df_BA.at[33:42,'Category']='Policy admin renewals'
    df_BA.at[43:62,'Category']='Finance & accounting'
    df_BA.at[63:80,'Category']='Claims'
    df_BA.at[81:85,'Category']='Payments hub (shared component)'
    df_BA.at[86:95,'Category']='Clients & contacts'
    df_BA.at[96:103,'Category']='Product building'
    df_BA.at[104,'Category']='Rating (not in the actual product)'
    df_BA.at[105:106,'Category']='Product migration'
    df_BA.at[107:108,'Category']='Front-end product building tool'
    df_BA.at[109,'Category']='Front-end configuration'
    df_BA.at[110:122,'Category']='Reinsurance'
    df_BA.at[123:126,'Category']='Reporting'
    df_BA.at[127:130,'Category']='Reporting tool & dashboard'
    df_BA.at[131:134,'Category']='Frontend processing'
    df_BA.at[135:139,'Category']='Work management'
    df_BA.at[140:154,'Category']='Document management'
    df_BA.at[155:161,'Category']='Kofax'
    df_BA.at[162:164,'Category']='Portfolio transfers'
    df_BA.at[165,'Category']='Workflow'
    df_BA.at[166,'Category']='DTU & migration tool'
    df_BA.at[167,'Category']='Batch processing (outside of the above categories)'
    df_BA.at[168:171,'Category']='Compliance'
elif businessAreasResult == 'Domain':
    #Create the data frame
    from itertools import islice
    ws = wb["Domain"]
    df_BA = load_workbook_range('A1:D172', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["D"].replace({"Basic": 1}, inplace=True)
    df_BA["D"].replace({"Intermediate": 2}, inplace=True)
    df_BA["D"].replace({"Advanced": 3}, inplace=True)
    df_BA["D"].replace({"SME": 4}, inplace=True)
    df_BA["D"].replace({"Skill Level": 'Skill Level'}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Description'})
    df_BA = df_BA.rename(columns={'D': 'Skill Level'})
    df_BA.at[1:29,'Category']='Insurance principles'
    df_BA.at[30:37,'Category']='Broking fundamentals'
    df_BA.at[38:82,'Category']='Underwriting'
    df_BA.at[83:106,'Category']='Claims'
    df_BA.at[107:129,'Category']='Finance'
    df_BA.at[130:150,'Category']='Reinsurance'
    df_BA.at[151:166,'Category']='Product'
    df_BA.at[167:170,'Category']='Territory'
    df_BA.at[171,'Category']='Qualifications'
elif businessAreasResult == 'BA':
    #Create the data frame
    from itertools import islice
    ws = wb["BA"]
    df_BA = load_workbook_range('A1:D94', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["D"].replace({"Basic": 1}, inplace=True)
    df_BA["D"].replace({"Intermediate": 2}, inplace=True)
    df_BA["D"].replace({"Advanced": 3}, inplace=True)
    df_BA["D"].replace({"SME": 4}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Description'})
    df_BA = df_BA.rename(columns={'D': 'Skill Level'})
    df_BA.at[1:9,'Category']='BA planning & monitoring'
    df_BA.at[10:24,'Category']='Elicitation & collaboration'
    df_BA.at[25:35,'Category']='Requirements management lifecycle'
    df_BA.at[36:48,'Category']='Strategy Analysis'
    df_BA.at[49:62,'Category']='Requirements analysis & design definition'
    df_BA.at[63:67,'Category']='Solution evaluation'
    df_BA.at[68:91,'Category']='Underlying competencies'
    df_BA.at[92:93,'Category']='Qualifications/Experience'
elif businessAreasResult == 'Architecture':
    #Create the data frame
    from itertools import islice
    ws = wb["Architecture"]
    df_BA = load_workbook_range('A1:D76', ws)

    # Create and Populate the score column based on answer provided
    df_BA = df_BA.fillna(value=0)
    df_BA["D"].replace({"": 0}, inplace=True)
    df_BA["D"].replace({"Basic": 1}, inplace=True)
    df_BA["D"].replace({"Intermediate": 2}, inplace=True)
    df_BA["D"].replace({"Advanced": 3}, inplace=True)
    df_BA["D"].replace({"SME": 4}, inplace=True)
    df_BA = df_BA.rename(columns={'A': 'Category'})
    df_BA = df_BA.rename(columns={'B': 'Topic'})
    df_BA = df_BA.rename(columns={'C': 'Description'})
    df_BA = df_BA.rename(columns={'D': 'Skill Level'})
    #delete first row that is filled with blank values
    #df_BA.drop(1, axis=0, inplace=True)
    # fill in category and topic labels
    df_BA.at[1:38,'Category']='Underlying competencies'
    df_BA.at[1:7,'Topic']='Analytical thinking & problem solving'
    df_BA.at[8:11,'Topic']='Communication skills'
    df_BA.at[12:15,'Topic']='Interaction skills'
    df_BA.at[16:23,'Topic']='Tools'
    df_BA.at[24:28,'Topic']='Architectural'
    df_BA.at[29:35,'Topic']='Process'
    df_BA.at[36:62,'Category']='Qualifications/Experience'
    df_BA.at[36:41,'Topic']='Architecture'
    df_BA.at[42:47,'Topic']='Programming'
    df_BA.at[49:53,'Topic']='Databases'
    df_BA.at[54:55,'Topic']='Methodology'
    df_BA.at[56:62,'Topic']='Leadership'
    df_BA.at[63:67,'Category']='Infrastructure'
    df_BA.at[63:64,'Topic']='Cloud'
    df_BA.at[65,'Topic']='Physical & virtual non cloud based'
    df_BA.at[66,'Topic']='Monitoring'
    df_BA.at[67:72,'Category']='Accountability'
    df_BA.at[73:75,'Category']='Preferences'

all_data = df_BA
all_data.fillna('', inplace=True)
print(all_data)

new_header = all_data.iloc[0] 
all_data = all_data[1:] 
all_data.columns = new_header
all_data['Skill Level'] = all_data['Skill Level'].astype('int')
print(all_data)

catList = all_data.Category.unique()
FilterCategories = (catList)
Catresult = st.sidebar.selectbox('Categories', FilterCategories)

dfCategory = all_data.groupby('Category')['Skill Level'].mean().sort_values(ascending=True)
dfCat = pd.DataFrame()
dfCat['Category'] = dfCategory.index
dfCat['Skill Level'] = dfCategory.values


chart = alt.Chart(dfCat).mark_bar().encode(alt.X('Category', sort=None,axis=alt.Axis(labelAngle=-45)), y='Skill Level').properties(title='Average Score by Category', width=800, height=600)
chart = chart.configure_title(
    fontSize=20,
    font='Helvetica',
    anchor='start',
    color='gray'
)
chart.save('Average score by category.html')

filterCat = all_data[all_data.Category.eq(Catresult)]
print(filterCat)

#Create chart to show individually selected category so you can see topic scores
chart2 = alt.Chart(filterCat).mark_bar().encode(
    alt.X('Topic', sort=None,axis=alt.Axis(labelAngle=-45)), 
    alt.Y('Skill Level', scale=alt.Scale(domain=[0,4]))).properties(title='Topic Score for Category - ' + Catresult, width=800, height=600)

chart2 = chart2.configure_title(
    fontSize=20,
    font='Helvetica',
    anchor='start',
    color='gray'
)
chart2.save('CategoryTopicScore.html')

#Create the chart show all topics and associated skill level
chart1 = alt.Chart(all_data).mark_bar().encode(alt.X('Topic', sort=None,axis=alt.Axis(labelAngle=-45)), y='Skill Level').properties(title='Score by Topic', width=1200, height=600)
chart1 = chart1.configure_title(
    fontSize=20,
    font='Helvetica',
    anchor='start',
    color='gray'
)
chart1.save('Scores by all topics.html')

#Create rest of streamlit main page including charts and overall scores
st.header(PersonName + ' - ' + businessAreasResult)
st.subheader('Overall results')
st.dataframe(df_Scores)
st.altair_chart(chart)
st.altair_chart(chart1)
st.altair_chart(chart2)