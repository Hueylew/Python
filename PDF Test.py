#!/usr/bin/env python3

import openpyxl
import os
import xlsxwriter
import pandas as pd
import numpy as np
import glob
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re
import fpdf
from fpdf import FPDF

f = "/Users/adamlewis/Documents/Work/Business Consultancy/Skills and Development/Combined Skills Profiles 2021/Adam Lewis - Skills Profile.xlsx"

from openpyxl import load_workbook
wb = load_workbook(f, data_only=True)
sh = wb["Details"]
PersonName = sh["B2"].value
PersonPos = sh["B3"].value
PersonScore = sh["J22"].value
PersonTarget = sh["J23"].value
PersonVariance = sh["J24"].value
PersonRole = sh["B3"].value
WorkIn = sh["B4"].value
Location = sh["B5"].value
YearsGI = sh["B6"].value
YearsBA = sh["B7"].value
YearsSSP = sh["B8"].value

document = fpdf.FPDF()

document.set_font('Helvetica', style='B', size=16)
document.set_text_color(19, 83, 173)
document.add_page()
document.cell(60)

document.cell(70, 10, PersonName, 1, 0, "C")
document.ln(2)

document.image('/Users/adamlewis/Library/Mobile Documents/com~apple~CloudDocs/Masons/26349018.png', 95, 22, 20)

document.ln(30)

document.set_font('Helvetica', style='', size=12)
document.set_text_color(0)
document.cell(0, h=5, txt = PersonRole)
document.ln

document.output('Report.pdf')
